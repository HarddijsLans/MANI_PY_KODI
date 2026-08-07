import csv
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


print("Programma uzsākta")


# ============================================================
# KONFIGURĀCIJA
# ============================================================

SAKOTNEJO_DATNU_MAPE = Path(
    r"C:\Users\hardijslans\Desktop\VISUAL STUDIO CODE"
    r"\VIRSRAKSTU PARVEIDE\ZEME ZEMES REFORMAS PABEIGSANAI"
    r"\Sakuma_datnes"
)

PUBLICESANAS_MAPE = Path(
    r"C:\Users\hardijslans\Desktop\VISUAL STUDIO CODE"
    r"\VIRSRAKSTU PARVEIDE\ZEME ZEMES REFORMAS PABEIGSANAI"
    r"\Datnes_publicesanai"
)


# ============================================================
# SĀKOTNĒJIE VIRSRĀKSTI
# ============================================================

SAKOTNEJIE_VIRSRAKSTI = [
    "Administratīvā teritorija",
    "Administratīvi teritoriālā vienība",
    "Zemes vienības kadastra apzīmējums",
    "Adrese",
    "Statuss (kods, nosaukums)",
    "Platība, ha",
    "Platība, m2",
    "Mežs, ha",
    "Mežs, m2",
    "Darbība, kuras rezultātā zemes vienība izveidota",
    "Kadastra apzīmējums zemes vienībai, no kuras izveidota zemes vienība",
    "Pazīme, ka uz zemes vienības ir reģistrētas būves",
    "Pazīme, ka uz zemes vienības ir reģistrētas būves ( būves, kam statuss nav pirmsreģistrēta)",
    "Pazīme, ka uz zemes vienības ir reģistrētas būves (tikai pirmsreģistrētās)",
    "Nekustamā īpašuma lietošanas mērķa kods un nosaukums",
    "Nekustamā īpašuma lietošanas mērķim piekrītošā platība, ha",
    "Nekustamā īpašuma lietošanas mērķim piekrītošā platība, m2",
    "Funkcionālās zonas lietošanas mērķu un neapbūvētas apbūves zemes atzīme",
    "Funkcionālās zonas lietošanas mērķu un neapbūvētas apbūves zemes atzīmes noteikšanas datums",
    "Atzīmes apraksts",
    "Atzīmes noteikšanas datums",
    "Atzīmes platība, ha",
    "Atzīmes platība, m2",
    "Fiskālā kadastrālā vērtība, Eur",
    "Fiskālās kadastrālās vērtības aprēķina datums",
    "Universālā kadastrālā vērtība, Eur",
    "Universālās kadastrālās vērtības aprēķina datums",
    "Mežaudzes vērtība, Eur",
    "VMD datu sagatavošanas datums",
    "Nekustamā īpašuma kadastra numurs",
    "Nekustamā īpašuma nosaukums",
    "Nekustamā īpašuma tiesiskais valdītājs",
    "Nekustamā īpašuma lietotājs",
    "Dati atlasīti uz",
]


# ============================================================
# JAUNIE CSV VIRSRĀKSTI
# ============================================================

JAUNIE_VIRSRAKSTI = [
    "AdmrKind",
    "AdmtKind",
    "ParCadNr",
    "Std",
    "ParLansId",
    "ParAreaHa",
    "ParAreaM",
    "ForestAreaHa",
    "ForestAreaM",
    "ParFormationType",
    "ParCadNrOld",
    "BuldReg",
    "RegBuilding",
    "PreRegBuilding",
    "PurlId",
    "PurlAreaHa",
    "PurlAreaM",
    "FunctionalZoneMark",
    "FunctionalZoneMarkDate",
    "MarkDescription",
    "MarkDate",
    "MarkAreaHa",
    "MarkAreaM",
    "FiscCadVal",
    "FiscCadValDate",
    "UnivCadVal",
    "UnivCadValDate",
    "ObjectForestValue",
    "ObjectForestValueDate",
    "ProCadNr",
    "ProName",
    "ProLegalHolder",
    "ProLegalPossessor",
    "Data",
]


# ============================================================
# PALĪGFUNKCIJA VIRSRĀKSTU PĀRBAUDEI
# ============================================================

def normalizet_virsrakstu(virsraksts):
    """
    Normalizē tikai virsraksta tekstu pārbaudei.

    Rindas pārnesumi, tabulācijas un vairākas atstarpes
    tiek aizstātas ar vienu atstarpi.

    Piemēram:
    VMD datu
    sagatavošanas datums

    kļūst par:
    VMD datu sagatavošanas datums
    """

    if virsraksts is None:
        return ""

    return " ".join(str(virsraksts).split())


# ============================================================
# SHĒMAS PĀRBAUDE
# ============================================================

print()
print("==============================")
print("SHĒMAS PĀRBAUDE")
print("==============================")


if len(SAKOTNEJIE_VIRSRAKSTI) != 34:
    print(
        f"❌ Kļūda kodā: sākotnējo virsrakstu skaits ir "
        f"{len(SAKOTNEJIE_VIRSRAKSTI)}, nevis 34."
    )
    sys.exit(1)


if len(JAUNIE_VIRSRAKSTI) != 34:
    print(
        f"❌ Kļūda kodā: jauno virsrakstu skaits ir "
        f"{len(JAUNIE_VIRSRAKSTI)}, nevis 34."
    )
    sys.exit(1)


print("✅ Ir 34 sākotnējie un 34 jaunie virsraksti.")


# ============================================================
# MAPJU PĀRBAUDE
# ============================================================

print()
print("==============================")
print("MAPJU PĀRBAUDE")
print("==============================")


if not SAKOTNEJO_DATNU_MAPE.exists():
    print("❌ Nav atrasta sākotnējo datņu mape:")
    print(SAKOTNEJO_DATNU_MAPE)
    sys.exit(1)


PUBLICESANAS_MAPE.mkdir(
    parents=True,
    exist_ok=True
)


print(f"📥 Ievades mape: {SAKOTNEJO_DATNU_MAPE}")
print(f"📤 Izvades mape: {PUBLICESANAS_MAPE}")


# ============================================================
# XLSX DATNES MEKLĒŠANA
# ============================================================

print()
print("==============================")
print("XLSX DATNES MEKLĒŠANA")
print("==============================")


xlsx_datnes = [
    datne
    for datne in SAKOTNEJO_DATNU_MAPE.glob("*.xlsx")
    if not datne.name.startswith("~$")
]


if len(xlsx_datnes) == 0:
    print("❌ Mapē nav atrasta neviena XLSX datne.")
    sys.exit(1)


if len(xlsx_datnes) > 1:
    print("❌ Mapē atrastas vairākas XLSX datnes.")
    print("Programma paredzēta tieši vienai XLSX datnei.")
    print()

    for datne in xlsx_datnes:
        print(f" - {datne.name}")

    sys.exit(1)


ievades_datne = xlsx_datnes[0]


# Gala XLSX saglabājam ar tādu pašu nosaukumu,
# bet mapē Datnes_publicesanai.

gala_xlsx = (
    PUBLICESANAS_MAPE
    / ievades_datne.name
)


# Gala CSV iegūst tādu pašu pamatnosaukumu.

gala_csv = (
    PUBLICESANAS_MAPE
    / ievades_datne.with_suffix(".csv").name
)


print(f"✅ Atrasta datne: {ievades_datne.name}")

print()
print(f"📄 Gala XLSX: {gala_xlsx.name}")
print(f"📄 Gala CSV : {gala_csv.name}")


# ============================================================
# ORIĢINĀLĀ XLSX ATVĒRŠANA
# ============================================================

print()
print("==============================")
print("XLSX NOLASĪŠANA")
print("==============================")


try:
    wb = load_workbook(ievades_datne)
except Exception as e:
    print(f"❌ Neizdevās atvērt XLSX datni: {e}")
    sys.exit(1)


ws = wb.active


print(f"✅ Darblapa: {ws.title}")
print(f"Rindu skaits: {ws.max_row}")
print(f"Kolonnu skaits: {ws.max_column}")


# ============================================================
# KOLONNU SKAITA PĀRBAUDE
# ============================================================

if ws.max_column != 34:

    print()
    print("❌ XLSX datnē nav 34 kolonnas.")
    print(f"Faktiski: {ws.max_column}")
    print("Paredzēts: 34")
    print()
    print("Datnes NETIKA pārveidotas.")

    sys.exit(1)


print("✅ Kolonnu skaits ir pareizs: 34.")


# ============================================================
# 3. RINDAS VIRSRĀKSTU PĀRBAUDE
# ============================================================

print()
print("==============================")
print("VIRSRĀKSTU PĀRBAUDE")
print("==============================")


faktiskie_virsraksti = [
    ws.cell(row=3, column=kolonna).value
    for kolonna in range(1, 35)
]


faktiskie_normalizeti = [
    normalizet_virsrakstu(v)
    for v in faktiskie_virsraksti
]


paredzetie_normalizeti = [
    normalizet_virsrakstu(v)
    for v in SAKOTNEJIE_VIRSRAKSTI
]


neatbilstibas = []


for nr, (faktiskais, paredzetais) in enumerate(
    zip(
        faktiskie_normalizeti,
        paredzetie_normalizeti
    ),
    start=1
):

    if faktiskais != paredzetais:

        neatbilstibas.append(
            (
                nr,
                faktiskais,
                paredzetais
            )
        )


if neatbilstibas:

    print("❌ XLSX virsraksti neatbilst paredzētajai shēmai.")
    print()

    for nr, faktiskais, paredzetais in neatbilstibas:

        print(f"Kolonna Nr. {nr}")
        print(f"  XLSX datnē : {repr(faktiskais)}")
        print(f"  Paredzēts   : {repr(paredzetais)}")
        print()

    print("Datnes NETIKA pārveidotas.")

    sys.exit(1)


print("✅ Visi 34 sākotnējie virsraksti atbilst shēmai.")


# ============================================================
# XLSX 1., 2. UN 4. RINDAS IZŅEMŠANA
# ============================================================

print()
print("==============================")
print("XLSX RINDU IZŅEMŠANA")
print("==============================")


# Dzēšam no apakšas uz augšu,
# lai rindu numuri dzēšanas laikā nenobīdītos.

ws.delete_rows(4, 1)
print("✅ Izņemta sākotnējā 4. rinda.")

ws.delete_rows(2, 1)
print("✅ Izņemta sākotnējā 2. rinda.")

ws.delete_rows(1, 1)
print("✅ Izņemta sākotnējā 1. rinda.")


# Pēc dzēšanas:
#
# sākotnējā 3. rinda -> jaunā 1. rinda
# sākotnējā 5. rinda -> jaunā 2. rinda


# ============================================================
# GALA XLSX PĀRBAUDE
# ============================================================

print()
print("==============================")
print("GALA XLSX PĀRBAUDE")
print("==============================")


gala_xlsx_virsraksti = [
    normalizet_virsrakstu(
        ws.cell(row=1, column=kolonna).value
    )
    for kolonna in range(1, 35)
]


if gala_xlsx_virsraksti != paredzetie_normalizeti:

    print(
        "❌ Pēc rindu dzēšanas XLSX pirmās rindas "
        "virsraksti nav pareizi."
    )

    sys.exit(1)


print("✅ XLSX 1. rindā ir 34 oriģinālie virsraksti.")
print("✅ XLSX dati sākas ar 2. rindu.")


# ============================================================
# GALA XLSX SAGLABĀŠANA
# ============================================================

print()
print("==============================")
print("XLSX SAGLABĀŠANA")
print("==============================")


try:

    wb.save(gala_xlsx)

except Exception as e:

    print(f"❌ Neizdevās saglabāt gala XLSX: {e}")
    sys.exit(1)


print(f"✅ XLSX saglabāts:")
print(gala_xlsx)


# ============================================================
# GALA XLSX NOLASĪŠANA CSV IZVEIDEI
# ============================================================

print()
print("==============================")
print("CSV SAGATAVOŠANA")
print("==============================")


try:

    df = pd.read_excel(
        gala_xlsx,
        header=0,
        dtype=object
    )

except Exception as e:

    print(
        f"❌ Neizdevās nolasīt gala XLSX CSV izveidei: {e}"
    )

    sys.exit(1)


if len(df.columns) != 34:

    print(
        f"❌ Gala XLSX kolonnu skaits nav 34: "
        f"{len(df.columns)}"
    )

    sys.exit(1)


# ============================================================
# CSV VIRSRĀKSTU NOMAIŅA
# ============================================================

df.columns = JAUNIE_VIRSRAKSTI


print("✅ CSV virsraksti nomainīti.")


# ============================================================
# CSV SAGLABĀŠANA
# ============================================================

try:

    df.to_csv(
        gala_csv,
        index=False,
        sep=",",
        encoding="utf-8-sig",
        quoting=csv.QUOTE_ALL,
        lineterminator="\n"
    )

except Exception as e:

    print(f"❌ Neizdevās saglabāt CSV datni: {e}")
    sys.exit(1)


print(f"✅ CSV saglabāts:")
print(gala_csv)


# ============================================================
# CSV VIRSRĀKSTU PĀRBAUDE
# ============================================================

print()
print("==============================")
print("GALA CSV PĀRBAUDE")
print("==============================")


try:

    with open(
        gala_csv,
        "r",
        encoding="utf-8-sig",
        newline=""
    ) as f:

        reader = csv.reader(
            f,
            delimiter=","
        )

        gala_csv_virsraksti = next(reader)

except Exception as e:

    print(f"❌ Neizdevās pārbaudīt gala CSV: {e}")
    sys.exit(1)


if gala_csv_virsraksti != JAUNIE_VIRSRAKSTI:

    print("❌ Gala CSV virsraksti nav pareizi.")
    sys.exit(1)


print("✅ Gala CSV satur visus 34 jaunos virsrakstus.")


# ============================================================
# PABEIGTS
# ============================================================

print()
print("==============================================")
print("PABEIGTS")
print("==============================================")
print()

print("📥 ORIĢINĀLĀ DATNE:")
print(ievades_datne)

print()

print("📗 GALA XLSX:")
print(gala_xlsx)

print()

print("📄 GALA CSV:")
print(gala_csv)

print()

print("✅ Oriģinālā XLSX datne nav mainīta.")
print("✅ No gala XLSX izņemta sākotnējā 1. rinda.")
print("✅ No gala XLSX izņemta sākotnējā 2. rinda.")
print("✅ No gala XLSX izņemta sākotnējā 4. rinda.")
print("✅ Oriģinālā 3. rinda kļuvusi par XLSX 1. rindu.")
print("✅ Oriģinālā 5. rinda kļuvusi par XLSX 2. rindu.")
print("✅ XLSX saglabāti oriģinālie latviešu virsraksti.")
print("✅ CSV izmantoti 34 jaunie virsraksti.")
print("✅ CSV atdalītājs ir komats.")
print("✅ CSV kodējums ir UTF-8 ar BOM.")
print("✅ CSV lauki tiek rakstīti pēdiņās.")

sys.exit(0)