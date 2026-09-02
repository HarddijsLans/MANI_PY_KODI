# ============================================================
# LADP publicējamie darījumi ar ZVB
# Versija: 1.0
#
# Funkcionalitāte:
# - meklē sākotnējo XLSX datni mapē SAKOTNEJAS DATNES
# - mapē drīkst būt tikai viena XLSX datne
# - datnes nosaukumam jāsākas ar:
#   ZVB_koordinatas_
# - prefiksa lielie/mazie burti nav būtiski
# - nolasa periodu DDMMYYYY-DDMMYYYY no datnes nosaukuma
# - validē perioda datumus
# - validē 10 gadu starpību starp sākuma un beigu gadskaitli
# - pārbauda, ka Excel darbgrāmatā ir tieši viena darba lapa
# - pārbauda precīzu 6 kolonnu struktūru un secību
# - pārveido tikai 1. rindas kolonnu virsrakstus
# - datu rindas nemaina
# - saglabā rezultātu mapē DATNES PUBLICESANAI
# - rezultāta nosaukumam pievieno _parveidota
# ============================================================

from datetime import datetime
from pathlib import Path
import sys

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# KONFIGURĀCIJA
# ============================================================

SAKOTNEJO_DATNU_MAPE = Path(
    r"C:\Users\hardijslans\Desktop\VISUAL STUDIO CODE"
    r"\VIRSRAKSTU PARVEIDE\DARIJUMI AR ZVB\SAKOTNEJAS DATNES"
)

PUBLICESANAS_MAPE = Path(
    r"C:\Users\hardijslans\Desktop\VISUAL STUDIO CODE"
    r"\VIRSRAKSTU PARVEIDE\DARIJUMI AR ZVB\DATNES PUBLICESANAI"
)

# Sākotnējās datnes nosaukumam jāsākas ar:
# ZVB_koordinatas_
#
# Salīdzināšana tiek veikta, neņemot vērā lielos/mazos burtus.
DATNES_PREFIKSS = "ZVB_koordinatas_"

DATUMA_FORMATS = "%d%m%Y"

PARVEIDOTAS_DATNES_PIELIKUMS = "_parveidota"

VIRSRAKSTU_RINDA = 1


# ============================================================
# SĀKOTNĒJIE KOLONNU VIRSRAKSTI
# ============================================================

SAKOTNEJIE_KOLONNU_NOSAUKUMI = [
    "BUICADNR",
    "DEAID",
    "KoordX",
    "KoordY",
    "DdN",
    "DdE",
]


# ============================================================
# PUBLICĒJAMIE KOLONNU NOSAUKUMI
# ============================================================

PUBLICEJAMIE_KOLONNU_NOSAUKUMI = [
    "Darījumā iekļautās būves kadastra apzīmējums.",
    (
        "Darījuma identifikators, kas izmantojams ieraksta "
        "sasaistīšanai ar darījuma datiem."
    ),
    (
        "Adresācijas objekta (būves)  centroīda X koordinātas, "
        "LKS-92 koordinātu sistēmā."
    ),
    (
        "Adresācijas objekta  (būves) centroīda Y koordinātas, "
        "LKS-92 koordinātu sistēmā."
    ),
    (
        "Adresācijas objekta  (būves) centroīda koordinātas "
        "platums (Latitude) decimālgrādos."
    ),
    (
        "Adresācijas objekta  (būves) centroīda koordinātas "
        "garums (Longitude) decimālgrādos."
    ),
]


# ============================================================
# DATNES NOSAUKUMA UN PERIODA VALIDĀCIJA
# ============================================================

def iegut_periodu_no_nosaukuma(
    datne: Path,
) -> tuple[datetime, datetime]:

    nosaukums = datne.stem

    # Pārbauda obligāto prefiksu.
    if not nosaukums.lower().startswith(
        DATNES_PREFIKSS.lower()
    ):
        raise ValueError(
            "Datnes nosaukums nesākas ar "
            f"'{DATNES_PREFIKSS}'."
        )

    # Noņem prefiksu un iegūst perioda daļu.
    perioda_teksts = nosaukums[len(DATNES_PREFIKSS):]

    perioda_dalas = perioda_teksts.split("-")

    if len(perioda_dalas) != 2:
        raise ValueError(
            "Pēc prefiksa jābūt periodam formātā "
            "DDMMYYYY-DDMMYYYY."
        )

    sakuma_datuma_teksts = perioda_dalas[0]
    beigu_datuma_teksts = perioda_dalas[1]

    if len(sakuma_datuma_teksts) != 8:
        raise ValueError(
            "Perioda sākuma datumam jābūt tieši 8 cipariem."
        )

    if len(beigu_datuma_teksts) != 8:
        raise ValueError(
            "Perioda beigu datumam jābūt tieši 8 cipariem."
        )

    if not sakuma_datuma_teksts.isdigit():
        raise ValueError(
            "Perioda sākuma datumā drīkst būt tikai cipari."
        )

    if not beigu_datuma_teksts.isdigit():
        raise ValueError(
            "Perioda beigu datumā drīkst būt tikai cipari."
        )

    try:
        sakuma_datums = datetime.strptime(
            sakuma_datuma_teksts,
            DATUMA_FORMATS,
        )

    except ValueError as kluda:
        raise ValueError(
            "Perioda sākuma datums nav korekts kalendāra datums."
        ) from kluda

    try:
        beigu_datums = datetime.strptime(
            beigu_datuma_teksts,
            DATUMA_FORMATS,
        )

    except ValueError as kluda:
        raise ValueError(
            "Perioda beigu datums nav korekts kalendāra datums."
        ) from kluda

    if sakuma_datums > beigu_datums:
        raise ValueError(
            "Perioda sākuma datums nedrīkst būt vēlāks "
            "par perioda beigu datumu."
        )

    if beigu_datums.year - sakuma_datums.year != 10:
        raise ValueError(
            "Perioda sākuma gada skaitlim jābūt tieši par "
            "10 mazākam nekā perioda beigu gada skaitlim."
        )

    return sakuma_datums, beigu_datums


# ============================================================
# SĀKOTNĒJĀS DATNES ATRAŠANA
# ============================================================

def atrast_sakotnejo_datni() -> Path:

    if not SAKOTNEJO_DATNU_MAPE.exists():
        raise FileNotFoundError(
            "Sākotnējo datņu mape neeksistē:\n"
            f"{SAKOTNEJO_DATNU_MAPE}"
        )

    xlsx_datnes = []

    for datne in SAKOTNEJO_DATNU_MAPE.glob("*.xlsx"):

        # Ignorē Excel pagaidu datnes.
        if datne.name.startswith("~$"):
            continue

        xlsx_datnes.append(datne)

    if len(xlsx_datnes) == 0:
        raise FileNotFoundError(
            "Mapē nav atrasta neviena XLSX datne.\n\n"
            "Sākotnējai datnei jāatrodas mapē:\n"
            f"{SAKOTNEJO_DATNU_MAPE}\n\n"
            "Datnes nosaukuma formāts:\n"
            "ZVB_koordinatas_DDMMYYYY-DDMMYYYY.xlsx\n\n"
            "Piemērs:\n"
            "ZVB_koordinatas_01012016-31082026.xlsx"
        )

    if len(xlsx_datnes) > 1:
        datnu_saraksts = "\n".join(
            f"  - {datne.name}"
            for datne in xlsx_datnes
        )

        raise ValueError(
            "Mapē SAKOTNEJAS DATNES drīkst būt tikai viena "
            "XLSX datne.\n\n"
            f"Atrasto XLSX datņu skaits: {len(xlsx_datnes)}\n\n"
            "Atrastas datnes:\n"
            f"{datnu_saraksts}"
        )

    sakotneja_datne = xlsx_datnes[0]

    # Validē datnes nosaukumu un periodu.
    iegut_periodu_no_nosaukuma(
        sakotneja_datne
    )

    return sakotneja_datne


# ============================================================
# REZULTĀTA DATNES CEĻŠ
# ============================================================

def izveidot_rezultata_celu(
    sakotneja_datne: Path,
) -> Path:

    rezultata_nosaukums = (
        f"{sakotneja_datne.stem}"
        f"{PARVEIDOTAS_DATNES_PIELIKUMS}.xlsx"
    )

    return PUBLICESANAS_MAPE / rezultata_nosaukums


# ============================================================
# DARBA LAPAS VALIDĀCIJA
# ============================================================

def parbaudit_darba_lapu_skaitu(
    darbgramata,
) -> None:

    lapu_skaits = len(darbgramata.worksheets)

    if lapu_skaits != 1:
        raise ValueError(
            "Excel darbgrāmatā jābūt tieši vienai darba lapai.\n\n"
            f"Faktiskais darba lapu skaits: {lapu_skaits}"
        )


# ============================================================
# KOLONNU VIRSRAKSTU NOLASĪŠANA
# ============================================================

def iegut_faktiskos_virsrakstus(
    darba_lapa,
) -> list:

    return [
        suna.value
        for suna in darba_lapa[VIRSRAKSTU_RINDA]
    ]


# ============================================================
# KOLONNU STRUKTŪRAS VALIDĀCIJA
# ============================================================

def parbaudit_kolonnu_strukturu(
    faktiskie_virsraksti: list,
) -> None:

    sagaidamais_kolonnu_skaits = len(
        SAKOTNEJIE_KOLONNU_NOSAUKUMI
    )

    faktiskais_kolonnu_skaits = len(
        faktiskie_virsraksti
    )

    if faktiskais_kolonnu_skaits != sagaidamais_kolonnu_skaits:
        raise ValueError(
            "Nepareizs kolonnu skaits.\n\n"
            f"Sagaidāmais kolonnu skaits: "
            f"{sagaidamais_kolonnu_skaits}\n"
            f"Faktiskais kolonnu skaits: "
            f"{faktiskais_kolonnu_skaits}"
        )

    if faktiskie_virsraksti == SAKOTNEJIE_KOLONNU_NOSAUKUMI:
        return

    kludu_rindas = []

    for indekss, faktiskais_nosaukums in enumerate(
        faktiskie_virsraksti
    ):

        sagaidamais_nosaukums = (
            SAKOTNEJIE_KOLONNU_NOSAUKUMI[indekss]
        )

        if faktiskais_nosaukums != sagaidamais_nosaukums:
            kludu_rindas.append(
                f"  - Kolonna {indekss + 1}: "
                f"atrasts '{faktiskais_nosaukums}', "
                f"sagaidāms '{sagaidamais_nosaukums}'"
            )

    kludu_teksts = "\n".join(
        kludu_rindas
    )

    raise ValueError(
        "Kolonnu struktūra neatbilst sagaidāmajai struktūrai.\n\n"
        "Kolonnu nosaukumiem un secībai jāsakrīt precīzi.\n\n"
        "Sagaidāmā secība:\n"
        "  BUICADNR\n"
        "  DEAID\n"
        "  KoordX\n"
        "  KoordY\n"
        "  DdN\n"
        "  DdE\n\n"
        "Konstatētās neatbilstības:\n"
        f"{kludu_teksts}"
    )


# ============================================================
# VIRSRAKSTU PĀRVEIDOŠANA
# ============================================================

def nomainit_lapas_virsrakstus(
    darba_lapa,
) -> int:

    faktiskie_virsraksti = iegut_faktiskos_virsrakstus(
        darba_lapa
    )

    parbaudit_kolonnu_strukturu(
        faktiskie_virsraksti
    )

    print()
    print("Kolonnu struktūra:")
    print("  Atpazīta korekta 6 kolonnu struktūra.")

    for indekss, jaunais_nosaukums in enumerate(
        PUBLICEJAMIE_KOLONNU_NOSAUKUMI,
        start=1,
    ):

        darba_lapa.cell(
            row=VIRSRAKSTU_RINDA,
            column=indekss,
        ).value = jaunais_nosaukums

    return len(
        PUBLICEJAMIE_KOLONNU_NOSAUKUMI
    )


# ============================================================
# DATNES PĀRVEIDOŠANA
# ============================================================

def parveidot_datni(
    sakotneja_datne: Path,
    rezultata_datne: Path,
) -> tuple[int, str]:

    try:
        darbgramata = load_workbook(
            filename=sakotneja_datne,
            data_only=False,
        )

    except PermissionError as kluda:
        raise PermissionError(
            "Sākotnējo datni nevar atvērt. "
            "Iespējams, tā pašlaik ir atvērta programmā Excel."
        ) from kluda

    except Exception as kluda:
        raise RuntimeError(
            "Neizdevās atvērt Excel datni.\n"
            f"Tehniskā informācija: {kluda}"
        ) from kluda

    try:
        parbaudit_darba_lapu_skaitu(
            darbgramata
        )

        darba_lapa = darbgramata.active
        lapas_nosaukums = darba_lapa.title

        parveidoto_skaits = nomainit_lapas_virsrakstus(
            darba_lapa
        )

        PUBLICESANAS_MAPE.mkdir(
            parents=True,
            exist_ok=True,
        )

        try:
            darbgramata.save(
                rezultata_datne
            )

        except PermissionError as kluda:
            raise PermissionError(
                "Rezultāta datni nevar saglabāt. "
                "Iespējams, datne ar šādu nosaukumu pašlaik "
                "ir atvērta programmā Excel."
            ) from kluda

        except Exception as kluda:
            raise RuntimeError(
                "Neizdevās saglabāt rezultāta datni.\n"
                f"Tehniskā informācija: {kluda}"
            ) from kluda

    finally:
        darbgramata.close()

    return parveidoto_skaits, lapas_nosaukums


# ============================================================
# GALVENĀ PROGRAMMA
# ============================================================

def main() -> None:

    print("=" * 70)
    print("ZVB DATNES KOLONNU VIRSRAKSTU PĀRVEIDOŠANA")
    print("=" * 70)

    sakotneja_datne = atrast_sakotnejo_datni()

    sakuma_datums, beigu_datums = (
        iegut_periodu_no_nosaukuma(
            sakotneja_datne
        )
    )

    rezultata_datne = izveidot_rezultata_celu(
        sakotneja_datne
    )

    print()
    print("Atrasta sākotnējā datne:")
    print(f"  {sakotneja_datne}")

    print()
    print("Datnes periods:")
    print(
        f"  {sakuma_datums.strftime('%d.%m.%Y')} - "
        f"{beigu_datums.strftime('%d.%m.%Y')}"
    )

    print()
    print("Rezultāta datne:")
    print(f"  {rezultata_datne}")

    if rezultata_datne.exists():
        print()
        print(
            "BRĪDINĀJUMS: rezultāta datne jau eksistē "
            "un tiks pārrakstīta."
        )

    parveidoto_skaits, lapas_nosaukums = parveidot_datni(
        sakotneja_datne=sakotneja_datne,
        rezultata_datne=rezultata_datne,
    )

    print()
    print("-" * 70)
    print("PĀRVEIDOŠANA VEIKSMĪGI PABEIGTA")
    print("-" * 70)

    print(
        f"Apstrādātā darba lapa: "
        f"{lapas_nosaukums}"
    )

    print(
        f"Nomainīto virsrakstu skaits: "
        f"{parveidoto_skaits}"
    )

    print()
    print(
        "Mainītas tikai 1. rindas "
        "6 virsrakstu šūnas."
    )
    print(
        "Datu rindas nav pārveidotas."
    )

    print()
    print("Pārveidotā datne saglabāta:")
    print(f"  {rezultata_datne}")

    print("=" * 70)


# ============================================================
# PROGRAMMAS PALAIŠANA
# ============================================================

if __name__ == "__main__":

    try:
        main()

    except Exception as kluda:

        print()
        print("=" * 70)
        print("KĻŪDA")
        print("=" * 70)
        print(str(kluda))
        print("=" * 70)

        sys.exit(1)