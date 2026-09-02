# ============================================================
# LADP publicējamie darījumi ar ZV
# Versija: 1.0
#
# Funkcionalitāte:
# - meklē sākotnējo XLSX datni mapē SAKOTNEJAS DATNES
# - datnes nosaukumam jāsākas ar:
#   ZV_vertibu_zonas_un_koordinatas_
# - nolasa periodu DDMMYYYY-DDMMYYYY no datnes nosaukuma
# - validē perioda datumus
# - validē 10 gadu starpību starp sākuma un beigu gadskaitli
# - ja ir vairākas derīgas datnes, izvēlas jaunāko pēc beigu datuma
# - pārbauda, ka Excel darbgrāmatā ir tieši viena darba lapa
# - pārbauda precīzu 11 kolonnu struktūru un secību
# - pieņem vienu no diviem pilniem virsrakstu variantiem
# - nepieļauj abu virsrakstu variantu jaukšanu
# - pārveido kolonnu virsrakstus uz publicējamiem nosaukumiem
# - datu rindas nemaina
# - saglabā rezultātu mapē DATNES PUBLICESANAI
# ============================================================

from datetime import datetime
from pathlib import Path
import sys

from openpyxl import load_workbook


# ============================================================
# KONFIGURĀCIJA
# ============================================================

SAKOTNEJO_DATNU_MAPE = Path(
    r"C:\Users\hardijslans\Desktop\VISUAL STUDIO CODE"
    r"\VIRSRAKSTU PARVEIDE\DARIJUMI AR ZV\SAKOTNEJAS DATNES"
)

PUBLICESANAS_MAPE = Path(
    r"C:\Users\hardijslans\Desktop\VISUAL STUDIO CODE"
    r"\VIRSRAKSTU PARVEIDE\DARIJUMI AR ZV\DATNES PUBLICESANAI"
)

# Sākotnējās datnes nosaukumam jāsākas ar:
# ZV_vertibu_zonas_un_koordinatas_
#
# Salīdzināšana tiek veikta, neņemot vērā lielos/mazos burtus.
DATNES_PREFIKSS = "ZV_vertibu_zonas_un_koordinatas_"

DATUMA_FORMATS = "%d%m%Y"

PARVEIDOTAS_DATNES_PIELIKUMS = "_parveidota"

VIRSRAKSTU_RINDA = 1


# ============================================================
# PIEĻAUJAMIE SĀKOTNĒJO KOLONNU VIRSRAKSTU VARIANTI
# ============================================================

# 1. variants
KOLONNU_VARIANTS_1 = [
    "PARCADNR",
    "DEAID",
    "VALZ_NAME1",
    "VALZ_NAME2",
    "VALZ_NAME3",
    "VALZ_NAME4",
    "VALZ_NAME5",
    "KoordX",
    "KoordY",
    "DdN",
    "DdE",
]

# 2. variants
KOLONNU_VARIANTS_2 = [
    "ParCadNr",
    "DeaId",
    "Valz1",
    "Valz2",
    "Valz3",
    "Valz4",
    "Valz5",
    "KoordX",
    "KoordY",
    "DdN",
    "DdE",
]


# ============================================================
# PUBLICĒJAMIE KOLONNU NOSAUKUMI
# ============================================================

PUBLICEJAMIE_KOLONNU_NOSAUKUMI = [
    "Darījumā iekļautās zemes vienības kadastra apzīmējums",
    "Darījuma identifikators",
    "Lauksaimniecībā izmantojamās zemes vērtību zona",
    "Meža zemes vērtību zonējums",
    "Dzīvojamo māju apbūves vērtību zona",
    "Rūpnieciskās ražošanas objektu apbūves vērtību zona",
    "Komercobjektu apbūves vērtību zona",
    "Adresācijas objekta centroīda X koordināta, LKS-92 koordinātu sistēmā",
    "Adresācijas objekta centroīda Y koordināta, LKS-92 koordinātu sistēmā",
    "Adresācijas objekta centroīda koordinātas platums (Latitude) decimālgrādos",
    "Adresācijas objekta centroīda koordinātas garums (Longitude) decimālgrādos",
]


# ============================================================
# DATNES NOSAUKUMA UN PERIODA VALIDĀCIJA
# ============================================================

def iegut_periodu_no_nosaukuma(
    datne: Path,
) -> tuple[datetime, datetime]:

    nosaukums = datne.stem

    # Pārbauda obligāto prefiksu.
    if not nosaukums.lower().startswith(
        DATNES_PREFIKSS.lower()
    ):
        raise ValueError(
            "Datnes nosaukums nesākas ar "
            f"'{DATNES_PREFIKSS}'."
        )

    # Noņem prefiksu un iegūst perioda daļu.
    perioda_teksts = nosaukums[len(DATNES_PREFIKSS):]

    perioda_dalas = perioda_teksts.split("-")

    if len(perioda_dalas) != 2:
        raise ValueError(
            "Pēc prefiksa jābūt periodam formātā "
            "DDMMYYYY-DDMMYYYY."
        )

    sakuma_datuma_teksts = perioda_dalas[0]
    beigu_datuma_teksts = perioda_dalas[1]

    if len(sakuma_datuma_teksts) != 8:
        raise ValueError(
            "Perioda sākuma datumam jābūt tieši 8 cipariem."
        )

    if len(beigu_datuma_teksts) != 8:
        raise ValueError(
            "Perioda beigu datumam jābūt tieši 8 cipariem."
        )

    if not sakuma_datuma_teksts.isdigit():
        raise ValueError(
            "Perioda sākuma datumā drīkst būt tikai cipari."
        )

    if not beigu_datuma_teksts.isdigit():
        raise ValueError(
            "Perioda beigu datumā drīkst būt tikai cipari."
        )

    try:
        sakuma_datums = datetime.strptime(
            sakuma_datuma_teksts,
            DATUMA_FORMATS,
        )
    except ValueError as kluda:
        raise ValueError(
            "Perioda sākuma datums nav korekts kalendāra datums."
        ) from kluda

    try:
        beigu_datums = datetime.strptime(
            beigu_datuma_teksts,
            DATUMA_FORMATS,
        )
    except ValueError as kluda:
        raise ValueError(
            "Perioda beigu datums nav korekts kalendāra datums."
        ) from kluda

    if sakuma_datums > beigu_datums:
        raise ValueError(
            "Perioda sākuma datums nedrīkst būt vēlāks "
            "par perioda beigu datumu."
        )

    if beigu_datums.year - sakuma_datums.year != 10:
        raise ValueError(
            "Perioda sākuma gada skaitlim jābūt tieši par "
            "10 mazākam nekā perioda beigu gada skaitlim."
        )

    return sakuma_datums, beigu_datums


# ============================================================
# SĀKOTNĒJĀS DATNES ATRAŠANA
# ============================================================

def atrast_jaunako_datni() -> Path:

    if not SAKOTNEJO_DATNU_MAPE.exists():
        raise FileNotFoundError(
            "Sākotnējo datņu mape neeksistē:\n"
            f"{SAKOTNEJO_DATNU_MAPE}"
        )

    kandidati = []

    for datne in SAKOTNEJO_DATNU_MAPE.glob("*.xlsx"):

        # Ignorē Excel pagaidu datnes.
        if datne.name.startswith("~$"):
            continue

        # Ignorē jau pārveidotas datnes.
        if datne.stem.lower().endswith(
            PARVEIDOTAS_DATNES_PIELIKUMS.lower()
        ):
            continue

        # Datnes nosaukumam obligāti jāsākas ar:
        # ZV_vertibu_zonas_un_koordinatas_
        if not datne.stem.lower().startswith(
            DATNES_PREFIKSS.lower()
        ):
            continue

        try:
            sakuma_datums, beigu_datums = (
                iegut_periodu_no_nosaukuma(datne)
            )

            kandidati.append(
                (
                    beigu_datums,
                    sakuma_datums,
                    datne,
                )
            )

        except ValueError as kluda:
            print()
            print("BRĪDINĀJUMS: izlaista neatbilstoša datne:")
            print(f"  {datne.name}")
            print(f"  Iemesls: {kluda}")

    if not kandidati:
        raise FileNotFoundError(
            "Netika atrasta neviena atbilstoša sākotnējā datne.\n\n"
            "Sākotnējai datnei jāatrodas mapē:\n"
            f"{SAKOTNEJO_DATNU_MAPE}\n\n"
            "Datnes nosaukumam jāsākas ar:\n"
            "ZV_vertibu_zonas_un_koordinatas_\n\n"
            "Pilnais nosaukuma formāts:\n"
            "ZV_vertibu_zonas_un_koordinatas_"
            "DDMMYYYY-DDMMYYYY.xlsx\n\n"
            "Piemērs:\n"
            "ZV_vertibu_zonas_un_koordinatas_"
            "01012016-31082026.xlsx"
        )

    # Izvēlas datni ar jaunāko perioda beigu datumu.
    _, _, jaunaka_datne = max(
        kandidati,
        key=lambda ieraksts: ieraksts[0],
    )

    return jaunaka_datne


# ============================================================
# REZULTĀTA DATNES CEĻŠ
# ============================================================

def izveidot_rezultata_celu(
    sakotneja_datne: Path,
) -> Path:

    rezultata_nosaukums = (
        f"{sakotneja_datne.stem}"
        f"{PARVEIDOTAS_DATNES_PIELIKUMS}.xlsx"
    )

    return PUBLICESANAS_MAPE / rezultata_nosaukums


# ============================================================
# DARBA LAPAS VALIDĀCIJA
# ============================================================

def parbaudit_darba_lapu_skaitu(darbgramata) -> None:

    lapu_skaits = len(darbgramata.worksheets)

    if lapu_skaits != 1:
        raise ValueError(
            "Excel darbgrāmatā jābūt tieši vienai darba lapai.\n\n"
            f"Faktiskais darba lapu skaits: {lapu_skaits}"
        )


# ============================================================
# KOLONNU VIRSRAKSTU NOLASĪŠANA
# ============================================================

def iegut_faktiskos_virsrakstus(darba_lapa) -> list:

    return [
        suna.value
        for suna in darba_lapa[VIRSRAKSTU_RINDA]
    ]


# ============================================================
# KOLONNU STRUKTŪRAS VALIDĀCIJA
# ============================================================

def parbaudit_kolonnu_strukturu(
    faktiskie_virsraksti: list,
) -> str:

    sagaidamais_kolonnu_skaits = len(KOLONNU_VARIANTS_1)
    faktiskais_kolonnu_skaits = len(faktiskie_virsraksti)

    if faktiskais_kolonnu_skaits != sagaidamais_kolonnu_skaits:
        raise ValueError(
            "Nepareizs kolonnu skaits.\n\n"
            f"Sagaidāmais kolonnu skaits: "
            f"{sagaidamais_kolonnu_skaits}\n"
            f"Faktiskais kolonnu skaits: "
            f"{faktiskais_kolonnu_skaits}"
        )

    # Pieļaujams pilns 1. variants.
    if faktiskie_virsraksti == KOLONNU_VARIANTS_1:
        return "1. variants"

    # Pieļaujams pilns 2. variants.
    if faktiskie_virsraksti == KOLONNU_VARIANTS_2:
        return "2. variants"

    # Ja nav precīzas atbilstības, nosaka tuvāko variantu,
    # lai kļūdas paziņojumā varētu parādīt neatbilstības.
    kludas_1 = []
    kludas_2 = []

    for indekss, faktiskais_nosaukums in enumerate(
        faktiskie_virsraksti
    ):

        sagaidamais_1 = KOLONNU_VARIANTS_1[indekss]
        sagaidamais_2 = KOLONNU_VARIANTS_2[indekss]

        if faktiskais_nosaukums != sagaidamais_1:
            kludas_1.append(
                (
                    indekss + 1,
                    faktiskais_nosaukums,
                    sagaidamais_1,
                )
            )

        if faktiskais_nosaukums != sagaidamais_2:
            kludas_2.append(
                (
                    indekss + 1,
                    faktiskais_nosaukums,
                    sagaidamais_2,
                )
            )

    if len(kludas_1) <= len(kludas_2):
        tuvakais_variants = "1. variants"
        kludas = kludas_1
    else:
        tuvakais_variants = "2. variants"
        kludas = kludas_2

    kludu_rindas = []

    for (
        kolonnas_numurs,
        faktiskais,
        sagaidamais,
    ) in kludas:

        kludu_rindas.append(
            f"  - Kolonna {kolonnas_numurs}: "
            f"atrasts '{faktiskais}', "
            f"sagaidāms '{sagaidamais}'"
        )

    kludu_teksts = "\n".join(kludu_rindas)

    raise ValueError(
        "Kolonnu struktūra neatbilst nevienam no "
        "pieļaujamajiem variantiem.\n\n"
        "Kolonnu nosaukumiem un secībai jāsakrīt ar vienu "
        "pilnu variantu.\n"
        "Abu variantu jaukšana nav pieļaujama.\n\n"
        f"Tuvākais pieļaujamais variants: {tuvakais_variants}\n\n"
        "Konstatētās neatbilstības:\n"
        f"{kludu_teksts}"
    )


# ============================================================
# VIRSRAKSTU PĀRVEIDOŠANA
# ============================================================

def nomainit_lapas_virsrakstus(darba_lapa) -> int:

    faktiskie_virsraksti = iegut_faktiskos_virsrakstus(
        darba_lapa
    )

    izmantotais_variants = parbaudit_kolonnu_strukturu(
        faktiskie_virsraksti
    )

    print()
    print("Kolonnu struktūra:")
    print(f"  Atpazīts {izmantotais_variants}")

    for indekss, jaunais_nosaukums in enumerate(
        PUBLICEJAMIE_KOLONNU_NOSAUKUMI,
        start=1,
    ):

        darba_lapa.cell(
            row=VIRSRAKSTU_RINDA,
            column=indekss,
        ).value = jaunais_nosaukums

    return len(PUBLICEJAMIE_KOLONNU_NOSAUKUMI)


# ============================================================
# DATNES PĀRVEIDOŠANA
# ============================================================

def parveidot_datni(
    sakotneja_datne: Path,
    rezultata_datne: Path,
) -> tuple[int, str]:

    try:
        darbgramata = load_workbook(
            filename=sakotneja_datne,
            data_only=False,
        )

    except PermissionError as kluda:
        raise PermissionError(
            "Sākotnējo datni nevar atvērt. "
            "Iespējams, tā pašlaik ir atvērta programmā Excel."
        ) from kluda

    except Exception as kluda:
        raise RuntimeError(
            "Neizdevās atvērt Excel datni.\n"
            f"Tehniskā informācija: {kluda}"
        ) from kluda

    try:
        parbaudit_darba_lapu_skaitu(darbgramata)

        darba_lapa = darbgramata.active
        lapas_nosaukums = darba_lapa.title

        parveidoto_skaits = nomainit_lapas_virsrakstus(
            darba_lapa
        )

        PUBLICESANAS_MAPE.mkdir(
            parents=True,
            exist_ok=True,
        )

        try:
            darbgramata.save(rezultata_datne)

        except PermissionError as kluda:
            raise PermissionError(
                "Rezultāta datni nevar saglabāt. "
                "Iespējams, datne ar šādu nosaukumu pašlaik "
                "ir atvērta programmā Excel."
            ) from kluda

        except Exception as kluda:
            raise RuntimeError(
                "Neizdevās saglabāt rezultāta datni.\n"
                f"Tehniskā informācija: {kluda}"
            ) from kluda

    finally:
        darbgramata.close()

    return parveidoto_skaits, lapas_nosaukums


# ============================================================
# GALVENĀ PROGRAMMA
# ============================================================

def main() -> None:

    print("=" * 70)
    print("ZV DATNES KOLONNU VIRSRAKSTU PĀRVEIDOŠANA")
    print("=" * 70)

    sakotneja_datne = atrast_jaunako_datni()

    sakuma_datums, beigu_datums = (
        iegut_periodu_no_nosaukuma(sakotneja_datne)
    )

    rezultata_datne = izveidot_rezultata_celu(
        sakotneja_datne
    )

    print()
    print("Atrasta sākotnējā datne:")
    print(f"  {sakotneja_datne}")

    print()
    print("Datnes periods:")
    print(
        f"  {sakuma_datums.strftime('%d.%m.%Y')} - "
        f"{beigu_datums.strftime('%d.%m.%Y')}"
    )

    print()
    print("Rezultāta datne:")
    print(f"  {rezultata_datne}")

    if rezultata_datne.exists():
        print()
        print(
            "BRĪDINĀJUMS: rezultāta datne jau eksistē "
            "un tiks pārrakstīta."
        )

    parveidoto_skaits, lapas_nosaukums = parveidot_datni(
        sakotneja_datne=sakotneja_datne,
        rezultata_datne=rezultata_datne,
    )

    print()
    print("-" * 70)
    print("PĀRVEIDOŠANA VEIKSMĪGI PABEIGTA")
    print("-" * 70)

    print(f"Apstrādātā darba lapa: {lapas_nosaukums}")
    print(f"Nomainīto virsrakstu skaits: {parveidoto_skaits}")

    print()
    print("Mainītas tikai virsrakstu rindas šūnas.")
    print("Datu rindas nav pārveidotas.")

    print()
    print("Pārveidotā datne saglabāta:")
    print(f"  {rezultata_datne}")

    print("=" * 70)


# ============================================================
# PROGRAMMAS PALAIŠANA
# ============================================================

if __name__ == "__main__":

    try:
        main()

    except Exception as kluda:

        print()
        print("=" * 70)
        print("KĻŪDA")
        print("=" * 70)
        print(str(kluda))
        print("=" * 70)

        sys.exit(1)